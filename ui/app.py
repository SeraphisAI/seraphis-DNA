from __future__ import annotations

import hashlib
import io
import json
import os
import time
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests
import streamlit as st

from katopu_core.omega_engine import apply_editspec, intent_to_editspec
from katopu_policy.evaluator import PolicyEvaluator
from katopu_shared.contract_utils import ERR_SERVER_ERROR, make_error_result
from katopu_shared.ids import CONTRACT_VERSION, ENGINE_VERSION, RESULT_SCHEMA_ID, SPEC_VERSION

# Optional export deps
try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as rl_canvas
except Exception:  # pragma: no cover
    A4 = None
    rl_canvas = None


# -----------------------------
# Config
# -----------------------------

DEFAULT_API_BASE = os.getenv("KATOPU_API_BASE", "http://localhost:8000").rstrip("/")
DEFAULT_TIMEOUT = float(os.getenv("KATOPU_HTTP_TIMEOUT", "8"))

POLICY_BASE_PATH = os.getenv("KATOPU_POLICY_PATH", "/policies/default.policy.json")
POLICY_OVERRIDE_PATH = os.getenv("KATOPU_POLICY_OVERRIDE_PATH", "/data/policy_override.json")
POLICY_AUDIT_PATH = os.getenv("KATOPU_POLICY_AUDIT_PATH", "/data/policy_audit.log")
TELEMETRY_PATH = os.getenv("KATOPU_TELEMETRY_PATH", "/data/telemetry.jsonl")

REPORT_DIR = os.getenv("KATOPU_REPORT_DIR", "/data/reports")
REPORT_MAX = int(os.getenv("KATOPU_REPORT_MAX", "20"))
REPORT_INDEX_PATH = os.path.join(REPORT_DIR, "index.json")


# -----------------------------
# Helpers
# -----------------------------

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def redact_seq(seq: str, head: int = 6, tail: int = 6) -> Dict[str, Any]:
    seq = seq or ""
    if len(seq) <= head + tail:
        preview = seq
    else:
        preview = f"{seq[:head]}…{seq[-tail:]}"
    return {"preview": preview, "len": len(seq), "sha256": sha256_hex(seq) if seq else None}


def tail_lines(path: str, n: int = 200) -> str:
    try:
        with open(path, "rb") as f:
            f.seek(0, os.SEEK_END)
            size = f.tell()
            block = 4096
            data = b""
            while size > 0 and data.count(b"\n") <= n:
                step = min(block, size)
                size -= step
                f.seek(size)
                data = f.read(step) + data
            text = data.decode("utf-8", errors="replace")
            lines = text.splitlines()[-n:]
            return "\n".join(lines)
    except FileNotFoundError:
        return "(dosya yok)"
    except Exception as e:
        return f"(okunamadı: {e})"


def ensure_report_dir() -> None:
    os.makedirs(REPORT_DIR, exist_ok=True)


def load_report_index() -> List[Dict[str, Any]]:
    ensure_report_dir()
    if not os.path.exists(REPORT_INDEX_PATH):
        return []
    try:
        obj = json.loads(open(REPORT_INDEX_PATH, "r", encoding="utf-8").read())
        if isinstance(obj, list):
            return obj
    except Exception:
        pass
    return []


def save_report_index(rows: List[Dict[str, Any]]) -> None:
    ensure_report_dir()
    with open(REPORT_INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(rows[:REPORT_MAX], f, ensure_ascii=False, indent=2)


def persist_report_json(report: Dict[str, Any]) -> str:
    """Persist report JSON and index entry; returns report_id."""
    ensure_report_dir()
    rid = report.get("id") or uuid.uuid4().hex
    report["id"] = rid
    report.setdefault("created_at", now_iso())

    # Save JSON
    path = os.path.join(REPORT_DIR, f"{rid}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # Update index
    idx = load_report_index()
    idx = [r for r in idx if r.get("id") != rid]
    idx.insert(
        0,
        {
            "id": rid,
            "created_at": report.get("created_at"),
            "name": report.get("name") or "Katopu Report",
            "intent": report.get("intent"),
            "mode": report.get("mode"),
            "source": report.get("source"),
            "effect_label": (report.get("result") or {}).get("effect_label"),
        },
    )
    save_report_index(idx)
    return rid


def load_report_json(report_id: str) -> Optional[Dict[str, Any]]:
    try:
        path = os.path.join(REPORT_DIR, f"{report_id}.json")
        if not os.path.exists(path):
            return None
        return json.loads(open(path, "r", encoding="utf-8").read())
    except Exception:
        return None


def common_affix_diff(before: str, after: str) -> Dict[str, str]:
    """Simple diff: common prefix/suffix, return middle parts."""
    b = before or ""
    a = after or ""
    i = 0
    m = min(len(b), len(a))
    while i < m and b[i] == a[i]:
        i += 1

    j = 0
    mb = len(b) - i
    ma = len(a) - i
    while j < min(mb, ma) and b[len(b) - 1 - j] == a[len(a) - 1 - j]:
        j += 1

    return {
        "prefix": b[:i],
        "before_mid": b[i : len(b) - j],
        "after_mid": a[i : len(a) - j],
        "suffix": b[len(b) - j :] if j > 0 else "",
    }


def report_json_bytes(report: Dict[str, Any]) -> bytes:
    return json.dumps(report, ensure_ascii=False, indent=2).encode("utf-8")


def report_xlsx_bytes(report: Dict[str, Any]) -> Optional[bytes]:
    if Workbook is None:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    def row(k: str, v: Any) -> None:
        ws.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)])

    row("id", report.get("id"))
    row("created_at", report.get("created_at"))
    row("name", report.get("name"))
    row("intent", report.get("intent"))
    row("mode", report.get("mode"))
    row("source", report.get("source"))

    seq = report.get("sequence") or {}
    row("sequence.before.len", (seq.get("before") or "").__len__())
    row("sequence.after.len", (seq.get("after") or "").__len__())

    res = report.get("result") or {}
    row("effect_label", res.get("effect_label"))
    row("metrics", res.get("metrics"))
    row("errors", res.get("errors"))

    # Spec sheet
    ws2 = wb.create_sheet("Spec")
    spec = report.get("spec")
    ws2.append(["spec_json"])
    ws2.append([json.dumps(spec, ensure_ascii=False, indent=2) if spec else "(none)"])

    # Result sheet
    ws3 = wb.create_sheet("Result")
    ws3.append(["result_json"])
    ws3.append([json.dumps(res, ensure_ascii=False, indent=2)])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def report_pdf_bytes(report: Dict[str, Any]) -> Optional[bytes]:
    if rl_canvas is None or A4 is None:
        return None

    bio = io.BytesIO()
    c = rl_canvas.Canvas(bio, pagesize=A4)
    width, height = A4

    y = height - 40

    def line(txt: str, dy: int = 14) -> None:
        nonlocal y
        c.drawString(40, y, txt[:120])
        y -= dy

    res = report.get("result") or {}
    seq = report.get("sequence") or {}

    line("Katopu GenLab — Report")
    line(f"ID: {report.get('id')}")
    line(f"Created: {report.get('created_at')}")
    line(f"Name: {report.get('name')}")
    line(f"Intent: {report.get('intent')}")
    line(f"Mode: {report.get('mode')} | Source: {report.get('source')}")
    line(f"Effect: {res.get('effect_label')} | Delta: {(res.get('metrics') or {}).get('delta_nt')}")

    line(" ")
    b = seq.get("before") or ""
    a = seq.get("after") or ""
    d = common_affix_diff(b, a)

    line("Before (preview): " + redact_seq(b).get("preview", ""))
    line("After  (preview): " + redact_seq(a).get("preview", ""))
    line(" ")

    line("Diff (middle parts):")
    line("Removed: " + (d["before_mid"][:80] if d["before_mid"] else "(none)"))
    line("Added:   " + (d["after_mid"][:80] if d["after_mid"] else "(none)"))

    line(" ")
    errs = res.get("errors") or []
    if errs:
        line("Errors:")
        for e in errs[:5]:
            line(f"- {e.get('code')}: {str(e.get('detail'))[:90]}")

    c.showPage()
    c.save()
    return bio.getvalue()


@dataclass
class RunResult:
    ok: bool
    source: str
    result: Optional[Dict[str, Any]] = None
    spec: Optional[Dict[str, Any]] = None
    error: Optional[str] = None


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"Accept": "application/json"})
    return s


def api_headers(api_key: str) -> Dict[str, str]:
    h: Dict[str, str] = {}
    if api_key.strip():
        h["X-API-Key"] = api_key.strip()
    return h


def api_post(session: requests.Session, api_base: str, path: str, payload: Dict[str, Any], timeout: float, api_key: str) -> Dict[str, Any]:
    url = f"{api_base}{path}"
    r = session.post(url, json=payload, headers=api_headers(api_key), timeout=timeout)
    # UI should prefer showing a stable body even when HTTP status is 4xx/5xx.
    try:
        data = r.json()
    except Exception:
        data = {"http_status": r.status_code, "detail": (r.text or "").strip()}
    return data


def api_get(session: requests.Session, api_base: str, path: str, timeout: float, api_key: str) -> Dict[str, Any]:
    url = f"{api_base}{path}"
    r = session.get(url, headers=api_headers(api_key), timeout=timeout)
    try:
        return r.json()
    except Exception:
        return {"http_status": r.status_code, "detail": (r.text or "").strip()}


def is_editspec(obj: Any) -> bool:
    return isinstance(obj, dict) and obj.get("schema") == "katopu.editspec.v1"


def is_result(obj: Any) -> bool:
    return isinstance(obj, dict) and obj.get("schema") == RESULT_SCHEMA_ID


def policy_active_path() -> Tuple[str, bool]:
    """Return (path, is_override)."""
    if os.path.exists(POLICY_OVERRIDE_PATH):
        return POLICY_OVERRIDE_PATH, True
    return POLICY_BASE_PATH, False


def get_policy_evaluator() -> Tuple[PolicyEvaluator, str, bool]:
    path, is_override = policy_active_path()
    key = "_policy_cache"
    cache = st.session_state.get(key)
    mtime = os.path.getmtime(path) if os.path.exists(path) else 0

    if cache is None or cache.get("path") != path or cache.get("mtime") != mtime:
        pe = PolicyEvaluator(path)
        st.session_state[key] = {"path": path, "mtime": mtime, "pe": pe, "override": is_override}
    cache = st.session_state[key]
    return cache["pe"], cache["path"], cache["override"]


def local_spec(intent: str, mode: str) -> Dict[str, Any]:
    """Create a contract-shaped EditSpec locally.

    This is used when the API is down *and* for demo-data generation.
    It must match katopu.editspec.v1 (shape-stable) to keep exports deterministic.
    """
    intent_norm = (intent or "").strip()
    strict_mode = str(mode or "strict").lower().strip() != "lenient"

    spec = intent_to_editspec(intent_norm, strict_mode=strict_mode)

    # Mirror API meta fields for easier debugging and stable exports.
    spec.setdefault("_meta", {})
    spec["_meta"].update(
        {
            "contract": CONTRACT_VERSION,
            "engine": ENGINE_VERSION,
            "source": "ui-local",
            "request_id": uuid.uuid4().hex,
        }
    )

    # Hard guard: never ship a spec without spec_version.
    if not spec.get("spec_version"):
        spec["spec_version"] = SPEC_VERSION

    return spec



def local_apply_with_policy(sequence: str, intent: str, mode: str) -> RunResult:
    """Local (API-free) execution path that still enforces PolicyEvaluator.evaluate()."""
    pe, pol_path, pol_override = get_policy_evaluator()
    strict_mode = str(mode or "strict").lower().strip() != "lenient"

    spec = local_spec(intent, "strict" if strict_mode else "lenient")
    request_id = (spec.get("_meta") or {}).get("request_id") or uuid.uuid4().hex

    try:
        result = apply_editspec(
            sequence=sequence,
            spec=spec,
            strict_mode=strict_mode,
            policy=pe,
            request_id=request_id,
            source="ui-local",
        )
    except Exception as exc:
        result = make_error_result(
            before=(sequence or ""),
            code=ERR_SERVER_ERROR,
            message="UI local engine error",
            detail={"type": type(exc).__name__, "msg": str(exc)},
            request_id=request_id,
            source="ui-local",
            spec_version=str(spec.get("spec_version", SPEC_VERSION)),
            policy_meta={"allowed": False, "reason": "ui-local-exception", "policy_id": None, "audit_id": None},
        )

    # UI wants to show which policy file is active.
    result.setdefault("_policy", {})
    result["_policy"].update({"policy_path": pol_path, "override": pol_override})

    ok = is_result(result) and not (result.get("errors") or [])
    err = None if ok else "local execution produced errors"
    return RunResult(ok=ok, source="local", result=result, spec=spec, error=err)


def run_chain(sequence: str, intent: str, mode: str, api_base: str, timeout: float, api_key: str, enable_local_fallback: bool) -> RunResult:
    session = st.session_state.get("_http")
    if session is None:
        session = make_session()
        st.session_state["_http"] = session

    # Prefer spec-first chain for determinism and better debuggability:
    # intent -> editspec -> apply -> result
    try:
        spec_payload = {"sequence": sequence, "intent": intent, "mode": mode}
        spec = api_post(session, api_base, "/nl/spec", spec_payload, timeout, api_key)
        if not is_editspec(spec):
            raise ValueError(f"invalid editspec: {spec}")

        apply_payload = {"sequence": sequence, "spec": spec}
        res = api_post(session, api_base, "/edit/apply", apply_payload, timeout, api_key)
        if not is_result(res):
            raise ValueError(f"invalid result: {res}")
        return RunResult(ok=True, source="api:/nl/spec+/edit/apply", result=res, spec=spec)
    except Exception as e:
        err1 = str(e)

    # Back-compat fallbacks: older API paths
    payload = {"intent": intent, "mode": mode}
    try:
        res = api_post(session, api_base, f"/run?sequence={sequence}", payload, timeout, api_key)
        if is_result(res):
            return RunResult(ok=True, source="api:/run", result=res)
        raise ValueError(f"invalid result: {res}")
    except Exception as e:
        err2 = str(e)

    try:
        res = api_post(session, api_base, f"/v1/katopu/run?sequence={sequence}", payload, timeout, api_key)
        if is_result(res):
            return RunResult(ok=True, source="api:/v1/katopu/run", result=res)
        raise ValueError(f"invalid result: {res}")
    except Exception as e:
        err3 = str(e)

    if not enable_local_fallback:
        return RunResult(ok=False, source="api", error=f"API failed. spec-chain: {err1} | /run: {err2} | /v1/katopu/run: {err3}")

    # 3) Local
    lr = local_apply_with_policy(sequence, intent, mode)
    if lr.ok:
        return lr
    lr.error = f"API failed; local fallback also failed: {lr.error}"
    return lr




# -----------------------------
# Sunuma Hazır Demo Modu Helpers
# -----------------------------

STATUS_CACHE_TTL_SEC = int(os.getenv('KATOPU_UI_STATUS_TTL_SEC', '10'))

def _ui_try_rerun() -> None:
    try:
        st.rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            pass


def get_status_snapshot(api_base: str, timeout: float, api_key: str) -> Dict[str, Any]:
    # Query /health and /policy with short caching; returns compact status for UI rozetleri.
    key = '_status_cache'
    now = time.time()
    cached = st.session_state.get(key)
    if isinstance(cached, dict) and (now - float(cached.get('ts', 0))) < STATUS_CACHE_TTL_SEC:
        return cached

    sess = st.session_state.get('_http') or make_session()
    st.session_state['_http'] = sess

    status: Dict[str, Any] = {
        'ts': now,
        'checked_at': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%SZ'),
        'api_ok': False,
        'policy_ok': False,
        'contract': CONTRACT_VERSION,
        'engine': ENGINE_VERSION,
        'policy_path': None,
        'error': None,
    }

    api_base = api_base.rstrip('/')

    try:
        h = api_get(sess, api_base, '/health', timeout, api_key)
        status['api_ok'] = True
        if isinstance(h, dict):
            meta = h.get('_meta') or {}
            if isinstance(meta, dict):
                status['contract'] = meta.get('contract') or status['contract']
                status['engine'] = meta.get('engine') or status['engine']
    except Exception as e:
        status['error'] = f"/health: {e}"

    try:
        pol = api_get(sess, api_base, '/policy', timeout, api_key)
        status['policy_ok'] = True
        if isinstance(pol, dict):
            status['policy_path'] = pol.get('path')
            meta = pol.get('_meta') or {}
            if isinstance(meta, dict):
                status['contract'] = meta.get('contract') or status['contract']
                status['engine'] = meta.get('engine') or status['engine']
    except Exception as e:
        if status.get('error'):
            status['error'] = status['error'] + f" | /policy: {e}"
        else:
            status['error'] = f"/policy: {e}"

    st.session_state[key] = status
    return status


def reset_reports_dir() -> int:
    # Delete report files (JSON + index) under REPORT_DIR; returns deleted file count.
    ensure_report_dir()
    deleted = 0
    try:
        for name in os.listdir(REPORT_DIR):
            path = os.path.join(REPORT_DIR, name)
            if os.path.isfile(path) and (name.endswith('.json') or name.endswith('.xlsx') or name.endswith('.pdf') or name.endswith('.zip')):
                try:
                    os.remove(path)
                    deleted += 1
                except Exception:
                    pass
        if os.path.exists(REPORT_INDEX_PATH):
            try:
                os.remove(REPORT_INDEX_PATH)
                deleted += 1
            except Exception:
                pass
    except Exception:
        pass
    return deleted


def _sequence_obj_for_report(before: str, after: str) -> Dict[str, Any]:
    if st.session_state.get('include_full_seq', True):
        return {'before': before, 'after': after}
    return {'before': '', 'after': '', 'before_redacted': redact_seq(before), 'after_redacted': redact_seq(after)}


def load_demo_dataset(n: int = 10, mode: str = "strict") -> Dict[str, Any]:
    # Create demo dataset (default 10) locally (no API required).
    deleted = reset_reports_dir()

    base_seq = 'ATGACCTTGGCTAACCTGTTACGATGGCCTTAA'
    demo_intents = [
        'ilk 5 baz sil',
        'son 5 baz sil',
        'başlangıçtan 3 baz çıkar',
        'sondan 2 baz sil',
        'ilk 1 baz sil',
        'son 1 baz sil',
        'ilk 2 baz sil',
        'son 2 baz sil',
        'ilk 0 baz sil',
        "pozisyon 5'deki bazı B yap",
    ]

    created = 0
    for i, intent in enumerate(demo_intents[: max(1, n)], start=1):
        rr = local_apply_with_policy(base_seq, intent, mode)
        spec = rr.spec or local_spec(intent, mode)
        res = rr.result if isinstance(rr.result, dict) else {}
        before = res.get('before') or base_seq
        after = res.get('after') or before

        report = {
            'id': uuid.uuid4().hex,
            'created_at': now_iso(),
            'name': f"DEMO #{i:02d} — {intent}",
            'intent': intent,
            'mode': mode,
            'source': rr.source,
            'sequence': _sequence_obj_for_report(before, after),
            'spec': spec,
            'result': res,
        }
        persist_report_json(report)
        created += 1

    return {'deleted': deleted, 'created': created}


def report_bundle_zip_bytes(report: Dict[str, Any]) -> bytes:
    # 1-click export bundle: PDF + XLSX + JSON in one ZIP.
    rid = (report.get('id') or 'report')
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode='w', compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"katopu_report_{rid}.json", report_json_bytes(report))

        pdf = report_pdf_bytes(report)
        if pdf is not None:
            z.writestr(f"katopu_report_{rid}.pdf", pdf)
        else:
            z.writestr('notes/MISSING_PDF.txt', 'PDF export requires reportlab.\n')

        xlsx = report_xlsx_bytes(report)
        if xlsx is not None:
            z.writestr(f"katopu_report_{rid}.xlsx", xlsx)
        else:
            z.writestr('notes/MISSING_XLSX.txt', 'Excel export requires openpyxl.\n')

        z.writestr(
            'ABOUT.txt',
            f"Generated: {now_iso()}\nContract: {CONTRACT_VERSION}\nEngine: {ENGINE_VERSION}\n",
        )

    return mem.getvalue()


def bundle_many_reports_zip_bytes(reports: List[Dict[str, Any]]) -> bytes:
    # ZIP many reports (each has JSON + optional PDF/XLSX).
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode='w', compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr(
            'ABOUT.txt',
            f"Generated: {now_iso()}\nCount: {len(reports)}\nContract: {CONTRACT_VERSION}\nEngine: {ENGINE_VERSION}\n",
        )
        for rep in reports:
            if not isinstance(rep, dict):
                continue
            rid = rep.get('id') or uuid.uuid4().hex
            z.writestr(f"{rid}/report.json", report_json_bytes(rep))
            px = report_pdf_bytes(rep)
            if px is not None:
                z.writestr(f"{rid}/report.pdf", px)
            xx = report_xlsx_bytes(rep)
            if xx is not None:
                z.writestr(f"{rid}/report.xlsx", xx)
    return mem.getvalue()


def ui_export_bundle_buttons(report: Dict[str, Any]) -> None:
    st.download_button(
        label='1‑Click Export Bundle (PDF+XLSX+JSON) — ZIP indir',
        data=report_bundle_zip_bytes(report),
        file_name=f"katopu_export_bundle_{report.get('id') or 'report'}.zip",
        mime='application/zip',
    )

# -----------------------------
# UI Rendering
# -----------------------------

st.set_page_config(page_title="Katopu GenLab (Ultra)", layout="wide")

st.title("Katopu GenLab — Ultra")
st.caption("Spec-first akış + policy panel + rapor tarihçesi + PDF/XLSX/JSON export")

# Sunuma hazır rozetler (API / Policy / Contract)
_api_base_for_status = st.session_state.get('api_base', DEFAULT_API_BASE).rstrip('/')
_api_key_for_status = st.session_state.get('api_key', '')
_timeout_for_status = float(st.session_state.get('timeout', DEFAULT_TIMEOUT))
_status = get_status_snapshot(_api_base_for_status, _timeout_for_status, _api_key_for_status)

st.markdown(
    "<style>" +
    ".katopu-badge {display:inline-block; padding:2px 10px; border-radius:999px; font-size:12px; margin-right:8px; border:1px solid rgba(0,0,0,0.15);}" +
    "</style>",
    unsafe_allow_html=True,
)

b1, b2, b3 = st.columns([1, 1, 2])
with b1:
    st.markdown(
        f"<span class='katopu-badge'>Status: {'API Healthy ✅' if _status.get('api_ok') else 'API Offline ⛔'}</span>",
        unsafe_allow_html=True,
    )
with b2:
    st.markdown(
        f"<span class='katopu-badge'>Policy: {'Active ✅' if _status.get('policy_ok') else 'Unknown ⛔'}</span>",
        unsafe_allow_html=True,
    )
with b3:
    st.markdown(
        f"<span class='katopu-badge'>Contract: {_status.get('contract')}</span> "
        f"<span class='katopu-badge'>Engine: {_status.get('engine')}</span> "
        f"<span class='katopu-badge'>Checked: {_status.get('checked_at')}</span>",
        unsafe_allow_html=True,
    )

if _status.get('policy_path'):
    st.caption(f"Policy path: {_status.get('policy_path')}")


with st.sidebar:
    st.subheader("Menü")
    page = st.radio(
        "",
        [
            "Lab",
            "DNA Önce/Sonra",
            "İndirilebilir Raporlar",
            "DNA Rapor Tarihçesi",
            "Batch",
            "Ultra Menüler",
            "Policy",
            "Telemetry",
            "Diagnostics",
        ],
        index=0,
    )

    st.markdown("---")
    st.subheader("Bağlantı")
    api_base = st.text_input("API Base", value=st.session_state.get("api_base", DEFAULT_API_BASE))
    api_key = st.text_input("API Key (opsiyonel)", value=st.session_state.get("api_key", ""), type="password")
    timeout = st.number_input(
        "HTTP Timeout (sn)",
        min_value=1.0,
        max_value=60.0,
        value=float(st.session_state.get("timeout", DEFAULT_TIMEOUT)),
        step=1.0,
    )

    enable_local_fallback = st.checkbox("API yoksa Local Fallback", value=bool(st.session_state.get("enable_local_fallback", True)))

    st.session_state["api_base"] = api_base.rstrip("/")
    st.session_state["api_key"] = api_key
    st.session_state["timeout"] = timeout
    st.session_state["enable_local_fallback"] = enable_local_fallback

    st.markdown("---")
    st.subheader("Rapor Ayarları")
    persist_reports = st.checkbox("Raporları kaydet (tarihçe)", value=bool(st.session_state.get("persist_reports", True)))
    include_full_seq = st.checkbox("Raporlarda tam DNA sakla", value=bool(st.session_state.get("include_full_seq", True)))
    st.session_state["persist_reports"] = persist_reports
    st.session_state["include_full_seq"] = include_full_seq

    st.markdown("---")
    st.subheader("Sunuma Hazır Demo Modu")
    st.caption("10 hazır rapor + 1-click export bundle + status rozetleri")
    c_demo1, c_demo2 = st.columns(2)
    with c_demo1:
        if st.button("Demo raporları yükle (10)"):
            info = load_demo_dataset(10, mode="strict")
            st.success(f"Demo dataset hazır: {info.get('created', 0)} rapor")
            _ui_try_rerun()
    with c_demo2:
        if st.button("Tarihçeyi temizle"):
            deleted = reset_reports_dir()
            st.success(f"Rapor klasörü temizlendi: {deleted} dosya")
            _ui_try_rerun()

    if st.button("Sağlık kontrolü (/health)"):

        session = st.session_state.get("_http") or make_session()
        st.session_state["_http"] = session
        try:
            r = api_get(session, api_base.rstrip("/"), "/health", timeout, api_key)
            st.success(r)
        except Exception as e:
            st.error(str(e))


def maybe_persist_run(
    *,
    sequence: str,
    intent: str,
    mode: str,
    rr: RunResult,
    report_name: str,
    spec: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    if not st.session_state.get("persist_reports", True):
        return None
    if not isinstance(rr.result, dict):
        return None

    # Best-effort: ensure before/after fields
    before = rr.result.get("before") or sequence
    after = rr.result.get("after") or before

    seq_obj: Dict[str, Any]
    if st.session_state.get("include_full_seq", True):
        seq_obj = {"before": before, "after": after}
    else:
        seq_obj = {"before": "", "after": "", "before_redacted": redact_seq(before), "after_redacted": redact_seq(after)}

    report = {
        "id": uuid.uuid4().hex,
        "created_at": now_iso(),
        "name": report_name.strip() or "Katopu Report",
        "intent": intent,
        "mode": mode,
        "source": rr.source,
        "sequence": seq_obj,
        "spec": spec,
        "result": rr.result,
    }

    rid = persist_report_json(report)
    st.session_state["last_report_id"] = rid
    return rid


def ui_download_buttons(report: Dict[str, Any]) -> None:
    c1, c2, c3 = st.columns(3)

    with c1:
        st.download_button(
            label="JSON indir",
            data=report_json_bytes(report),
            file_name=f"katopu_report_{report.get('id')}.json",
            mime="application/json",
        )

    with c2:
        x = report_xlsx_bytes(report)
        if x is None:
            st.info("Excel export için openpyxl gerekli.")
        else:
            st.download_button(
                label="Excel indir (.xlsx)",
                data=x,
                file_name=f"katopu_report_{report.get('id')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    with c3:
        p = report_pdf_bytes(report)
        if p is None:
            st.info("PDF export için reportlab gerekli.")
        else:
            st.download_button(
                label="PDF indir",
                data=p,
                file_name=f"katopu_report_{report.get('id')}.pdf",
                mime="application/pdf",
            )


def select_report_ui(idx: List[Dict[str, Any]], label: str = "Rapor seç") -> Optional[str]:
    if not idx:
        st.info("Henüz rapor yok.")
        return None

    def fmt(r: Dict[str, Any]) -> str:
        dt = r.get("created_at", "")
        rid = r.get("id", "")
        intent = r.get("intent", "")
        return f"{dt} | {rid[:8]} | {intent}".strip()

    options = {fmt(r): r.get("id") for r in idx}
    # Preselect last report if exists
    default = st.session_state.get("last_report_id")
    keys = list(options.keys())
    default_ix = 0
    if default:
        for i, k in enumerate(keys):
            if options[k] == default:
                default_ix = i
                break

    chosen = st.selectbox(label, keys, index=default_ix)
    return options.get(chosen)


# -----------------------------
# Pages
# -----------------------------

if page == "Lab":
    st.subheader("Lab")

    c1, c2 = st.columns([2, 1])

    with c1:
        sequence = st.text_area("DNA/RNA Sequence", value="ATGACCTTGGCTAACCTGTTACGATGGCCTTAA", height=120)
        intent = st.text_input("Intent (TR serbest metin)", value="ilk 5 baz sil")
        report_name = st.text_input("Rapor adı (opsiyonel)", value=st.session_state.get("report_name", ""))
        st.session_state["report_name"] = report_name

    with c2:
        mode = st.selectbox("Mode", ["strict", "lenient"], index=0)
        run_btn = st.button("Tek Adım Çalıştır", type="primary")
        spec_btn = st.button("Spec Üret")
        apply_btn = st.button("Spec'i Uygula")

    spec_store_key = "_spec_current"

    if spec_btn:
        sess = st.session_state.get("_http") or make_session()
        st.session_state["_http"] = sess
        payload = {"sequence": sequence, "intent": intent, "mode": mode}
        try:
            spec = api_post(sess, api_base, "/nl/spec", payload, timeout, api_key)
            st.session_state[spec_store_key] = spec
            st.success("Spec üretildi (API)")
        except Exception:
            spec = local_spec(intent, mode)
            st.session_state[spec_store_key] = spec
            st.warning("API spec alınamadı; local spec üretildi")

    if apply_btn:
        spec = st.session_state.get(spec_store_key)
        if not spec:
            st.error("Önce Spec Üret")
        else:
            sess = st.session_state.get("_http") or make_session()
            st.session_state["_http"] = sess
            try:
                # /edit/apply contract expects the full EditSpec (katopu.editspec.v1) under "spec".
                # Sending only an "edit" fragment results in schema=NULL / missing spec_version/op errors.
                payload = {"sequence": sequence, "spec": spec}
                res = api_post(sess, api_base, "/edit/apply", payload, timeout, api_key)
                rr = RunResult(ok=is_result(res), source="api:/edit/apply", result=res, spec=spec)
                st.success("Uygulandı (API)")
                st.json(res)
                rid = maybe_persist_run(sequence=sequence, intent=intent, mode=mode, rr=rr, report_name=report_name, spec=spec)
                if rid:
                    st.toast(f"Rapor kaydedildi: {rid[:8]}")
            except Exception:
                rr = local_apply_with_policy(sequence, intent, mode)
                if rr.ok:
                    st.success("Uygulandı (Local)")
                    st.json(rr.result)
                    rid = maybe_persist_run(sequence=sequence, intent=intent, mode=mode, rr=rr, report_name=report_name, spec=rr.spec)
                    if rid:
                        st.toast(f"Rapor kaydedildi: {rid[:8]}")
                else:
                    st.error(rr.error or "Local apply failed")
                    if rr.result:
                        st.json(rr.result)

    if run_btn:
        rr = run_chain(sequence, intent, mode, api_base, timeout, api_key, enable_local_fallback)
        if rr.ok:
            st.success(f"OK — source: {rr.source}")
            st.json(rr.result)
            if rr.spec:
                st.session_state[spec_store_key] = rr.spec
            rid = maybe_persist_run(sequence=sequence, intent=intent, mode=mode, rr=rr, report_name=report_name, spec=rr.spec or st.session_state.get(spec_store_key))
            if rid:
                st.toast(f"Rapor kaydedildi: {rid[:8]}")
        else:
            st.error(rr.error or "Run failed")
            if rr.result:
                st.json(rr.result)

    spec = st.session_state.get(spec_store_key)
    if spec:
        st.subheader("Current Spec")
        st.json(spec)

    # Quick access to last report
    idx = load_report_index()
    if idx:
        st.markdown("---")
        st.subheader("Son Rapor")
        rid = st.session_state.get("last_report_id") or idx[0].get("id")
        rep = load_report_json(rid) if rid else None
        if rep:
            ui_download_buttons(rep)
            ui_export_bundle_buttons(rep)


elif page == "DNA Önce/Sonra":
    st.subheader("DNA Önce / Sonra")
    idx = load_report_index()
    rid = select_report_ui(idx)
    if rid:
        rep = load_report_json(rid)
        if not rep:
            st.error("Rapor bulunamadı")
        else:
            seq = rep.get("sequence") or {}
            before = seq.get("before") or ""
            after = seq.get("after") or ""

            if not before and "before_redacted" in seq:
                st.warning("Bu rapor redacted saklanmış; tam DNA yok.")

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("### Önce")
                st.code(before if before else (seq.get("before_redacted") or {}).get("preview", ""))
            with c2:
                st.markdown("### Sonra")
                st.code(after if after else (seq.get("after_redacted") or {}).get("preview", ""))

            st.markdown("---")
            d = common_affix_diff(before, after)
            st.markdown("### Fark Özeti")
            cc1, cc2 = st.columns(2)
            with cc1:
                st.write("**Kaldırılan (removed)**")
                st.code(d["before_mid"] if d["before_mid"] else "(yok)")
            with cc2:
                st.write("**Eklenen (added)**")
                st.code(d["after_mid"] if d["after_mid"] else "(yok)")

            st.markdown("---")
            ui_download_buttons(rep)
            ui_export_bundle_buttons(rep)


elif page == "İndirilebilir Raporlar":
    st.subheader("İndirilebilir Raporlar")
    st.caption("PDF / Excel / JSON")

    idx = load_report_index()
    rid = select_report_ui(idx)
    if rid:
        rep = load_report_json(rid)
        if rep:
            ui_download_buttons(rep)
            ui_export_bundle_buttons(rep)
            st.markdown("---")
            st.subheader("Rapor İçeriği")
            st.json(rep)


elif page == "DNA Rapor Tarihçesi":
    st.subheader("DNA Rapor Tarihçesi")
    st.markdown('### 1‑Click Toplu Export (Son 10 Rapor)')
    if 'bundle_last10' not in st.session_state:
        st.session_state['bundle_last10'] = None

    cB1, cB2 = st.columns([1, 2])
    with cB1:
        if st.button('Son 10 raporu paketle (ZIP hazırla)'):
            reps = []
            for r in (idx or [])[:10]:
                rid0 = r.get('id')
                rep0 = load_report_json(rid0) if rid0 else None
                if rep0:
                    reps.append(rep0)
            st.session_state['bundle_last10'] = bundle_many_reports_zip_bytes(reps) if reps else None

    with cB2:
        if st.session_state.get('bundle_last10'):
            st.download_button(
                'Son 10 raporu indir (ZIP)',
                data=st.session_state['bundle_last10'],
                file_name=f"katopu_last10_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}Z.zip",
                mime='application/zip',
            )


    idx = load_report_index()

    if idx:
        # table
        rows = []
        for r in idx:
            rows.append(
                {
                    "created_at": r.get("created_at"),
                    "id": r.get("id"),
                    "name": r.get("name"),
                    "intent": r.get("intent"),
                    "mode": r.get("mode"),
                    "source": r.get("source"),
                    "effect": r.get("effect_label"),
                }
            )
        st.dataframe(rows, use_container_width=True, hide_index=True)

        st.markdown("---")
        rid = select_report_ui(idx, label="Detay için rapor seç")
        if rid:
            rep = load_report_json(rid)
            if rep:
                ui_download_buttons(rep)
                ui_export_bundle_buttons(rep)
                st.markdown("---")
                st.subheader("Görüntüle")
                st.json(rep)

                c1, c2 = st.columns([1, 3])
                with c1:
                    if st.button("Bu raporu sil"):
                        try:
                            path = os.path.join(REPORT_DIR, f"{rid}.json")
                            if os.path.exists(path):
                                os.remove(path)
                            idx2 = [x for x in load_report_index() if x.get("id") != rid]
                            save_report_index(idx2)
                            st.success("Silindi")
                        except Exception as e:
                            st.error(str(e))
                with c2:
                    st.caption("Silme işlemi yalnızca yerel tarihçeyi etkiler.")
    else:
        st.info("Henüz rapor yok. Lab sayfasında çalıştırınca otomatik kaydolur.")


elif page == "Batch":
    st.subheader("Batch")
    st.caption("Her satır: (sequence\tintent) veya (sequence,intent). İstersen intent-only modunu Ultra Menüler'den kullan.")

    default_batch = "ATGC\tilk 2 baz sil\nATGACCTTGGCTAACCTGTTACGATGGCCTTAA\tilk 5 baz sil"
    batch_text = st.text_area("Batch input", value=st.session_state.get("batch_text", default_batch), height=160)
    st.session_state["batch_text"] = batch_text

    mode_b = st.selectbox("Mode (batch)", ["strict", "lenient"], index=0, key="mode_batch")

    colA, colB, colC = st.columns([1, 1, 2])

    with colA:
        go = st.button("Batch Çalıştır", type="primary")
    with colB:
        redact_export = st.checkbox("Export redacted (default)", value=True)
    with colC:
        st.write("Export: redacted = sequence preview + sha256 (tam dizi yok)")

    def parse_lines(txt: str) -> List[Tuple[str, str]]:
        out: List[Tuple[str, str]] = []
        for raw in txt.splitlines():
            line = raw.strip()
            if not line:
                continue
            if line.startswith("#"):
                continue
            if "\t" in line:
                a, b = line.split("\t", 1)
            elif "," in line:
                a, b = line.split(",", 1)
            else:
                # sequence only; intent missing
                out.append((line, ""))
                continue
            out.append((a.strip(), b.strip()))
        return out

    if go:
        rows = parse_lines(batch_text)
        results: List[Dict[str, Any]] = []
        for seq, it in rows:
            rr = run_chain(seq, it, mode_b, api_base, timeout, api_key, enable_local_fallback)
            item: Dict[str, Any] = {
                "ok": rr.ok,
                "source": rr.source,
                "sequence": seq,
                "intent": it,
                "error": rr.error,
                "result": rr.result,
            }
            results.append(item)

        st.session_state["batch_results"] = results
        st.success(f"Bitti: {len(results)}")

    results = st.session_state.get("batch_results")
    if results:
        st.subheader("Results")
        st.json(results)

        export_obj: Dict[str, Any]
        if redact_export:
            export_rows: List[Dict[str, Any]] = []
            for r in results:
                rr = r.get("result") or {}
                before = rr.get("before") or ""
                after = rr.get("after") or ""
                export_rows.append(
                    {
                        "ok": r.get("ok"),
                        "source": r.get("source"),
                        "intent": r.get("intent"),
                        "before": redact_seq(before),
                        "after": redact_seq(after),
                        "effect_label": rr.get("effect_label"),
                        "edit_map": rr.get("edit_map"),
                        "metrics": rr.get("metrics"),
                        "_policy": rr.get("_policy"),
                        "_meta": rr.get("_meta"),
                        "error": r.get("error"),
                    }
                )
            export_obj = {"schema": "katopu.batch.export.redacted.v1", "rows": export_rows}
        else:
            export_obj = {"schema": "katopu.batch.export.full.v1", "rows": results}

        st.download_button(
            label="Download batch.json",
            data=json.dumps(export_obj, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name="katopu_batch.json",
            mime="application/json",
        )


elif page == "Ultra Menüler":
    st.subheader("Ultra Menüler")

    st.markdown("### 1) Contract Inspector")
    sess = st.session_state.get("_http") or make_session()
    st.session_state["_http"] = sess

    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("/policy oku"):
            try:
                p = api_get(sess, api_base, "/policy", timeout, api_key)
                st.success("OK")
                st.json(p)
            except Exception as e:
                st.error(str(e))
    with col2:
        if st.button("/openapi.json oku"):
            try:
                o = api_get(sess, api_base, "/openapi.json", timeout, api_key)
                st.success("OK")
                st.json({"title": o.get("info", {}).get("title"), "version": o.get("info", {}).get("version")})
            except Exception as e:
                st.error(str(e))

    st.markdown("---")
    st.markdown("### 2) Rate-Limit Torture (kontrollü)")
    st.caption("Bu test, aynı anda çoklu istek gönderir. API rate-limit'ini gözlemlemek için.")

    seq_t = st.text_input("Torture sequence", value="ATGACCTTGGCTAACCTGTTACGATGGCCTTAA")
    intents_t = st.text_area(
        "Intent list (her satır bir intent — # yorum satırı SKIP)",
        value="# smoke\nilk 5 baz sil\nson 2 baz sil\n# mutate (opsiyonel)\npozisyon 10'deki bazı G yap",
        height=120,
    )
    concurrency = st.slider("Concurrency", min_value=1, max_value=50, value=20)
    total = st.slider("Toplam istek (satırlar döngülenir)", min_value=1, max_value=300, value=60)

    mode_t = st.selectbox("Mode (torture)", ["strict", "lenient"], index=0, key="mode_torture")

    def _intent_lines(txt: str) -> List[str]:
        lines = []
        for r in txt.splitlines():
            s = r.strip()
            if not s or s.startswith("#"):
                continue
            lines.append(s)
        return lines

    if st.button("Torture Çalıştır"):
        from concurrent.futures import ThreadPoolExecutor, as_completed

        intents_list = _intent_lines(intents_t)
        if not intents_list:
            st.error("En az 1 intent gerekli")
        else:
            session = make_session()

            def task(i: int) -> Tuple[int, str]:
                intent_i = intents_list[i % len(intents_list)]
                payload = {"intent": intent_i, "mode": mode_t}
                try:
                    r = session.post(
                        f"{api_base}/run?sequence={seq_t}",
                        json=payload,
                        headers=api_headers(api_key),
                        timeout=timeout,
                    )
                    return r.status_code, intent_i
                except Exception:
                    return 0, intent_i

            counts: Dict[int, int] = {}
            t0 = time.time()
            with ThreadPoolExecutor(max_workers=concurrency) as ex:
                futs = [ex.submit(task, i) for i in range(total)]
                for f in as_completed(futs):
                    code, _ = f.result()
                    counts[code] = counts.get(code, 0) + 1
            dt = time.time() - t0

            st.success(f"Bitti: {total} istek / {dt:.2f}s")
            st.write("Status dağılımı:")
            st.json(counts)
            if 429 in counts:
                st.warning("429 görüldü (rate limit). Bu durumda body'nin result.v1 shape kaldığını ayrıca API tarafında doğrula.")

    st.markdown("---")
    st.markdown("### 3) Intent-only Batch (UCP batch) — tek sequence + satır satır intent")
    seq_b = st.text_input("Batch sequence", value="ATGACCTTGGCTAACCTGTTACGATGGCCTTAA", key="batch_seq")
    intents_b = st.text_area(
        "Batch intents (each line is one intent)",
        value="ilk 5 baz sil\nson 2 baz sil\n# yorum satırı\nilk 0 baz sil",
        height=140,
        key="batch_intents_only",
    )
    if st.button("Intent-only Batch Çalıştır"):
        lines = _intent_lines(intents_b)
        results: List[Dict[str, Any]] = []
        for it in lines:
            rr = run_chain(seq_b, it, mode_t, api_base, timeout, api_key, enable_local_fallback)
            results.append({"intent": it, "ok": rr.ok, "source": rr.source, "error": rr.error, "result": rr.result})
        st.session_state["intent_only_results"] = results
        st.success(f"Bitti: {len(results)}")

    r = st.session_state.get("intent_only_results")
    if r:
        st.json(r)
        st.download_button(
            "Download intent_only_batch.json",
            data=json.dumps({"schema": "katopu.batchrun.v1", "rows": r}, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name="katopu_intent_only_batch.json",
            mime="application/json",
        )


elif page == "Policy":
    st.subheader("Policy Panel")

    base_path = POLICY_BASE_PATH
    active_path, is_override = policy_active_path()

    c1, c2 = st.columns(2)

    with c1:
        st.caption(f"Base policy: {base_path}")
        st.code(tail_lines(base_path, n=200) if os.path.exists(base_path) else "(base policy yok)")

    with c2:
        st.caption(f"Active policy: {active_path} ({'override' if is_override else 'base'})")
        if os.path.exists(active_path):
            active_txt = open(active_path, "r", encoding="utf-8").read()
        else:
            active_txt = "{}"

        edited = st.text_area("Override editor (JSON)", value=active_txt, height=220)

        colX, colY, colZ = st.columns([1, 1, 2])
        with colX:
            if st.button("Override kaydet"):
                try:
                    obj = json.loads(edited)
                    if not isinstance(obj, dict):
                        raise ValueError("policy must be an object")
                    os.makedirs(os.path.dirname(POLICY_OVERRIDE_PATH), exist_ok=True)
                    with open(POLICY_OVERRIDE_PATH, "w", encoding="utf-8") as f:
                        json.dump(obj, f, ensure_ascii=False, indent=2)
                    st.success(f"Yazıldı: {POLICY_OVERRIDE_PATH}")
                except Exception as e:
                    st.error(str(e))

        with colY:
            if st.button("Override sıfırla"):
                try:
                    if os.path.exists(POLICY_OVERRIDE_PATH):
                        os.remove(POLICY_OVERRIDE_PATH)
                    st.success("Override kaldırıldı")
                except Exception as e:
                    st.error(str(e))

        with colZ:
            st.caption("Not: Override yazınca API otomatik olarak override policy'yi yükler (container restart gerekmez).")

    st.markdown("---")
    st.caption(f"Policy audit tail: {POLICY_AUDIT_PATH}")
    st.code(tail_lines(POLICY_AUDIT_PATH, n=120))


elif page == "Telemetry":
    st.subheader("Telemetry")
    st.caption("Telemetry default kapalıdır. Açmak için infra/.env içinde KATOPU_TELEMETRY_ENABLED=true yap.")
    st.caption(f"Telemetry tail: {TELEMETRY_PATH}")
    st.code(tail_lines(TELEMETRY_PATH, n=120))


elif page == "Diagnostics":
    st.subheader("Diagnostics")
    st.write("**UI**: http://localhost:8501")
    st.write("**API**: http://localhost:8000")
    st.write("**Swagger**: http://localhost:8000/docs")

    st.markdown("### PowerShell hızlı komutlar")
    st.code(
        """
# Proje klasörüne gir
cd "C:\\Users\\LENOVO\\Desktop\\katopu_genlab_ultra_v2\\infra"

# Başlat / güncelle
 docker compose up -d --build

# Durum
 docker ps --format "table {{.Names}}\t{{.Ports}}\t{{.Status}}"

# Loglar
 docker compose logs -f api
 docker compose logs -f ui

# Durdur
 docker compose down
""".strip(),
        language="powershell",
    )

